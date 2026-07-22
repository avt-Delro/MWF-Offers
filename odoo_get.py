from collections import defaultdict
import xmlrpc.client
from xmlrpc.client import ServerProxy, SafeTransport
import math
import ssl
from datetime import datetime
import calendar

import environment as env
from dotenv import load_dotenv
import os
from logging_config import logger

import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from copy import copy
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.styles import PatternFill, Border, Protection, Side
import traceback
from format_excel import config_ps_mb, mapping_excel


api_key = env.odoo_api
username = env.odoo_username
product_file = env.product_file
strapping_file = env.strapping_file
load_dotenv()



datetoday = datetime.now().strftime("%Y-%m-%d")

'''
Odoo Credential
'''
ODOO_URL_B2B = 'https://wtdus.odoo.com'
ODOO_DB_B2B = 'wtdus-production-19540213'
ODOO_USERNAME_B2B = env.odoo_username
ODOO_PASSWORD_B2B = env.odoo_api

ODOO_URL_B2B_jetstar = 'https://jetstar-trailer.odoo.com'
ODOO_DB_B2B_jetstar = 'jetstar-trailer-production-30682357'
ODOO_USERNAME_B2B_jetstar = env.odoo_username_jetstar
ODOO_PASSWORD_B2B_jetstar = env.odoo_api_jetstar

# context = ssl._create_unverified_context()
 
# common = xmlrpc.client.ServerProxy(
#     f"{ODOO_URL_B2B}/xmlrpc/2/common",
#     context=context
# )

# uid = common.authenticate(ODOO_DB_B2B, ODOO_USERNAME_B2B, ODOO_PASSWORD_B2B, {})

# if isinstance(uid, int) and uid > 0:
#     logger.info(f"Odoo login as ID: {uid}")
#     print('hello b2b')
# else:
#     logger.error("Failed")

# models = xmlrpc.client.ServerProxy(f"{ODOO_URL_B2B}/xmlrpc/2/object", context=context)

_uid = None
_models = None

_uid_2 = None
_models_2 = None


def get_models():
    global _uid, _models

    if _models is None:
        context = ssl._create_unverified_context()

        common = xmlrpc.client.ServerProxy(
            f"{ODOO_URL_B2B}/xmlrpc/2/common",
            context=context
        )

        _uid = common.authenticate(
            ODOO_DB_B2B,
            ODOO_USERNAME_B2B,
            ODOO_PASSWORD_B2B,
            {}
        )

        if isinstance(_uid, int) and _uid > 0:
            logger.info(f"Odoo login as ID: {_uid}")
        else:
            logger.error("Failed Odoo login")

        _models = xmlrpc.client.ServerProxy(
            f"{ODOO_URL_B2B}/xmlrpc/2/object",
            context=context
        )

    return _uid, _models

def get_models_v2():
    global _uid_2, _models_2

    if _models_2 is None:
        context = ssl._create_unverified_context()

        common = xmlrpc.client.ServerProxy(
            f"{ODOO_URL_B2B_jetstar}/xmlrpc/2/common",
            context=context
        )

        _uid_2 = common.authenticate(
            ODOO_DB_B2B_jetstar,
            ODOO_USERNAME_B2B_jetstar,
            ODOO_PASSWORD_B2B_jetstar,
            {}
        )

        if isinstance(_uid, int) and _uid > 0:
            logger.info(f"Odoo login as ID: {_uid}")
        else:
            logger.error("Failed Odoo login")

        _models_2 = xmlrpc.client.ServerProxy(
            f"{ODOO_URL_B2B_jetstar}/xmlrpc/2/object",
            context=context
        )

    return _uid_2, _models_2

def create_sheet(filepath, data_row, sheetname):
    if isinstance(data_row, list):
        df = pd.DataFrame(data_row)
    elif isinstance(data_row, dict):
        df = pd.DataFrame([data_row])
    
    with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer,  sheet_name=sheetname, index=False)
        logger.info(f"Data written to {sheetname} sheet in {filepath}")


def normalize(text: str) -> str:
    if not text:
        return ""
    return (
        text.lower()
        .replace('\u00a0', ' ')   # non-breaking space
        .replace('–', '-')        # en dash
        .replace('—', '-')        # em dash
        .replace('‑', '-')        # non-breaking hyphen
        .strip()
    )




#Checks if product is allowed
#def is_product_allowed(config_lookup, main_df):
#     if config_lookup['NAME'] != 'Point S Tire':
#         arroyo = config_lookup['ARROYO']
#         ars = config_lookup['ARS']
#         lcr = config_lookup['LCR']
#         tbr = config_lookup['TBR']
#         ars_tbr = config_lookup['ARS-TBR']
#         aryo_tbr = config_lookup['ARYO-TBR']
#         aryo_st = config_lookup['ARYO-ST']
#         st = config_lookup['ST']


#         main_df['FLAG'] = 'N'

#         brand_conditions = [
#             (arroyo == 'Y') & (main_df['BRAND'] == 'ARROYO'),
#             (ars == 'Y') & (main_df['BRAND'] == 'AMERICANROADSTAR'),
#         ]

#         brand_choices = ['Y', 'Y']

#         flag = np.select(brand_conditions, brand_choices, default='Y')

#         #Implicate type conditions
#         type_conditions = [
#             (lcr == 'Y') & (main_df['TYPE'] == 'LCR'),
#             (tbr == 'Y') & (main_df['TYPE'] == 'TBR'),

#             (ars_tbr == 'Y') & (main_df['TYPE'] == 'ARS-TBR'),
#             (ars_tbr == 'N') & (main_df['TYPE'] == 'ARS-TBR'),

#             (aryo_tbr == 'Y') & (main_df['TYPE'] == 'ARYO-TBR'),
#             (aryo_tbr == 'N') & (main_df['TYPE'] == 'ARYO-TBR'),

#             (aryo_st == 'Y') & (main_df['TYPE'] == 'ARYO-ST'),
#             (aryo_st == 'N') & (main_df['TYPE'] == 'ARYO-ST'),

#             (st == 'Y') & (main_df['TYPE'] == 'ST'),
#         ]

#         #Type Results
#         type_choices = [
#             'Y', 'Y',
#             'Y', 'N',
#             'Y', 'N',
#             'Y', 'N',
#             'Y',
#         ]
#         #Re-add config, with the initial flag value as a default.
#         main_df['FLAG'] = np.select(type_conditions, type_choices, default=flag)

#         filtered_df = main_df.loc[(main_df['FLAG'] == 'Y') & (main_df['WEIGHTS'] != 'None')]

#     else:
#         filtered_df = main_df[main_df['BRAND'].isin(config_ps.point_s_allowed_brand)]
    
#     return filtered_df



def is_product_allowed(config, df):

    if config['NAME'] == 'POINTSV2':
        return df[df['BRAND'].isin(config_ps_mb.point_s_allowed_brand)]
    
    if config['MB'] == 'N':
        #~ means negate or not
        df = df[~df['BRAND'].isin(config_ps_mb.mb_allowed_brand)]
    else:
        pass

    df = df.copy()

    # # --------------------------
    # # BRAND RULES
    # # Only restrict listed brands
    # # Others = allowed by default
    # # --------------------------
    # brand_rules = {
    #     'ARROYO': config['ARROYO'],
    #     'AMERICAN ROADSTAR': config['ARS'],
    # }

    # df['BRAND_OK'] = df['BRAND'].map(brand_rules).fillna('Y')  # ✅ FIX

    # --------------------------
    # TYPE RULES
    # --------------------------
    # type_rules = {
    #     'LCR': config['LCR'],
    #     'TBR': config['TBR'],

    #     'ARS-TBR': config['ARS-TBR'],
    #     'ARYO-TBR': config['ARYO-TBR'],
    #     'ARYO-ST': config['ARYO-ST'],
    #     'ST': config['ST'],
    # }

    type_rules = {
        'ARYO PCR': config['ARYO PCR'],
        'ARYO ST - GS': config['ARYO ST - GS'],
        'ARYO-TBR': config['ARYO-TBR'],
        'ARYO - OTHER': config['ARYO - OTHER'],
        'ARS PCR': config['ARS PCR'],
        'ARS ST - GS': config['ARS ST - GS'],
        'SAFFIRO ST - COFO': config['SAFFIRO ST - COFO'],
        'ARS ST - COFO': config['ARS ST - COFO'],
        'TBR - OTHER': config['TBR - OTHER'],
        'ST - OTHER': config['ST - OTHER'],
        'LCR': config['LCR'],
    }

    df['TYPE_OK'] = df['TYPE'].map(type_rules).fillna('Y')  # ✅ optional default allow

    # --------------------------
    # FINAL FLAG
    # --------------------------
    df['FLAG'] = (df['TYPE_OK'] == 'Y').map({True: 'Y', False: 'N'})

    # --------------------------
    # FINAL FILTER
    # --------------------------
    return df.loc[
        (df['FLAG'] == 'Y') &
        (df['WEIGHTS'] != 'None')
    ]

#Returns df that filters allowed product
def identify_prod_to_add(file, config, customer):
    products_df = pd.read_excel(file, sheet_name='Sheet2')
        #Normalize column names
    products_df.columns = products_df.columns.str.strip().str.upper()
    
    try:
       df = is_product_allowed(config, products_df)   

       
       logger.info(f"Product filtering completed for customer '{customer}'")
       
       return df
    except Exception as e:
        logger.error(f"Error processing config for customer '{customer}': {e}")
        return

#Return the last number from search size since it is used in filtering
def get_last_number(tire_size):
    
    if not tire_size:   # ✅ handles None, '', False
            return None

    tire_size = str(tire_size)


    match = re.search(r'[RD](?:F)?(\d+(?:\.\d+)?)', tire_size)
    if match:
        return match.group(1) 
    else:
        new_size = tire_size[-2:]
        return new_size

def groupby_je(filepath_exc, column_names, name_of_sheet, get_column):
    df = pd.read_excel(filepath_exc, engine='openpyxl', sheet_name=name_of_sheet)
    
    if isinstance(column_names, str):
        column_names = [column_names]


    code_dictionary = (df.groupby(column_names, dropna = False )[get_column]
                       .sum(min_count=1)
                       .fillna(0)
                       .reset_index()
                       .to_dict(orient = 'records')
                       )
    
    create_sheet(filepath_exc, code_dictionary, name_of_sheet)

#default_code:: List, loc_key:: String
def get_on_hand_quant(file, default_code, loc_key, sheet_name, odoo_instance):
    if odoo_instance == 1:
        uid, models = get_models()
        
        domain = [('product_reference_code', 'in', default_code)] + (
                [('location_id', 'ilike', loc_key)]
                if not isinstance(loc_key, list)
                else ['|'] * (len(loc_key) - 1) + [('location_id', 'ilike', k) for k in loc_key]
            )
        
        product_quant = models.execute_kw(ODOO_DB_B2B, uid, ODOO_PASSWORD_B2B, 'stock.quant', 'search_read', [domain],{
                'fields': ['product_reference_code', 'available_quantity', 'location_id']
            })
        
        
        create_sheet(file, product_quant, sheet_name)
    elif odoo_instance == 2:
        uid, models = get_models_v2()
        
        domain = [('product_id.product_tmpl_id.legacy_uid', 'in', default_code)] + (
                [('location_id', 'ilike', loc_key)]
                if not isinstance(loc_key, list)
                else ['|'] * (len(loc_key) - 1) + [('location_id', 'ilike', k) for k in loc_key]
            )
        
        product_quant = models.execute_kw(ODOO_DB_B2B_jetstar, uid, ODOO_PASSWORD_B2B_jetstar, 'stock.quant', 'search_read', [domain],{
                'fields': ['product_id','product_reference_code', 'available_quantity', 'location_id']
            })
        # Get unique product IDs
        product_ids = list({q['product_id'][0] for q in product_quant if q.get('product_id')})
    
        if product_ids:
            # Read products to get template IDs
            products = models.execute_kw(ODOO_DB_B2B_jetstar,uid,ODOO_PASSWORD_B2B_jetstar,'product.product','read',[product_ids],{'fields': ['product_tmpl_id']})
    
            template_ids = list({
                p['product_tmpl_id'][0]
                for p in products
                if p.get('product_tmpl_id')
            })
        
            # Read templates to get legacy_uid
            templates = models.execute_kw(ODOO_DB_B2B_jetstar,uid,ODOO_PASSWORD_B2B_jetstar,'product.template','read',[template_ids],{'fields': ['legacy_uid']})
        
            # template_id -> legacy_uid
            template_map = {
                t['id']: t['legacy_uid']
                for t in templates
            }
        
            # product_id -> legacy_uid
            product_map = {
                p['id']: template_map.get(p['product_tmpl_id'][0])
                for p in products
            }
        
            # Add part number to each quant
            for quant in product_quant:
                quant['part_number'] = product_map.get(
                quant['product_id'][0]
                )
        
        logger.info(f'Product quant: {type(product_quant)}: {product_quant}')
        
        create_sheet(file, product_quant, sheet_name)
        


def df_to_qty_map(df, product_col, qty_col='available_quantity'):
    #Filter rows for inventory that is greater or equal to 8
    # capped_qty = df[df[qty_col]>= 8].copy()

    # #For inventory that exceed 200, if >200 return 200
    # capped_qty[qty_col] = capped_qty[qty_col].clip(upper=200)
    # return dict(zip(capped_qty[product_col], capped_qty[qty_col]))
    if df.empty:
        return {}

    return dict(zip(df[product_col], df[qty_col]))
    
#Get odoo file one time run
def get_odoo_data(file):
    uid, models = get_models()
    try:
        strap_df = pd.read_excel(strapping_file, engine='openpyxl')

        strap_df['fullsize'] = strap_df['fullsize'].astype(str).str.strip().str.upper()

        '''
        Creates a lookup map for strapping
            {
                '4.1/3.50R4': 2.0,
                '4.8/4R8': 2.0,
                '8/3R4': 3.0,
                '3/8R4': 4.0
            }
        '''
        strap_map = dict(zip(strap_df['fullsize'], strap_df['weights']))


        #Dataframe from the main product file and append the result to Sheet2 
        df = pd.read_excel(file, sheet_name='Sheet1')
        df.columns = df.columns.str.strip().str.upper()
        logger.info("Detected columns: %s", df.columns.tolist())

        records = df.to_dict(orient='records')
        codes = df['PRODUCT'].dropna().unique().tolist()
        partnum = df['PN'].dropna().unique().tolist()
        logger.info(f"Unique default_codes extracted: {codes}")
        logger.info("Collected %d unique default_codes", len(codes))


        list_to_append_data = []
        taxes = models.execute_kw(ODOO_DB_B2B, uid, ODOO_PASSWORD_B2B, 'account.tax', 'search_read', [[]], {'fields':['id','amount']})
        tax_map = {tax['id']: tax['amount'] for tax in taxes}

        #Odoo domain filtering
        products_info = models.execute_kw(ODOO_DB_B2B, uid, ODOO_PASSWORD_B2B, 'product.template', 'search_read',[[('default_code', 'in', codes)]],{
                    'fields': ['default_code','full_size', 'brand', 'tire_class','tire_type', 'ply', 'model', 'load_index', 'speed_rating', 'lt_p', 'taxes_id', 'part', 'weight'],
                })
        
        #Replace all false values with -
        for product in products_info:
            if product.get('taxes_id') in [False, '0', 0]:
                product['taxes_id'] = '0'
                
            for key, values in product.items():
                if key != 'taxes_id' and values in [False, '0', 0, '-']:
                    product[key] = ''


        
        logger.info("Fetched %d products from Odoo", len(products_info))

        #Pomona
        get_on_hand_quant(file, codes, ['BIN326', '326/Input','326/PORT','326/WHL', '326/OLT'], 'Pomona', 1)
        groupby_je(file, 'product_reference_code', 'Pomona', 'available_quantity')

        #Carrolton
        get_on_hand_quant(file, codes, '331/Stock', 'Carrolton',1)

        #Latrobe
        get_on_hand_quant(file, codes, '381/Stock' , 'Latrobe',1)

        #Kentucky
        get_on_hand_quant(file, codes, '357/Stock', 'Kentucky',1)
        
        #Fort Worth
        get_on_hand_quant(file, codes, '380/Stock', 'Fort Worth',1)
        
        #Apopka
        get_on_hand_quant(file, codes, '321/Stock', 'Apopka',1)
        
        #Winschester
        get_on_hand_quant(file, codes, '376/Stock', 'Winchester',1)

        #Hileaeah - Miami
        get_on_hand_quant(file, codes, '305/Stock', 'Miami',1)

        #Bloomsburg
        get_on_hand_quant(file, codes, '570/Stock', 'Bloomsburg',1)

        #Winchester
        get_on_hand_quant(file, partnum, 'Package/', 'Tennesse', 2)

        pomona_quant = pd.read_excel(file, sheet_name='Pomona')
        carr_quant = pd.read_excel(file, sheet_name='Carrolton')
        latrobe_quant = pd.read_excel(file, sheet_name='Latrobe')
        kent_quant = pd.read_excel(file, sheet_name='Kentucky')
        fortworth_quant = pd.read_excel(file, sheet_name='Fort Worth')
        apopka_quant = pd.read_excel(file, sheet_name='Apopka')
        winchester_quant = pd.read_excel(file, sheet_name='Winchester')
        miami_quant = pd.read_excel(file, sheet_name='Miami')
        bloomsburg_quant = pd.read_excel(file, sheet_name='Bloomsburg')
        tennesse_quant = pd.read_excel(file, sheet_name='Tennesse')

        pm_map = df_to_qty_map(pomona_quant,'product_reference_code')
        carr_map = df_to_qty_map(carr_quant,'product_reference_code')
        latrobe_map = df_to_qty_map(latrobe_quant,'product_reference_code')
        kent_map = df_to_qty_map(kent_quant,'product_reference_code')
        fort_map = df_to_qty_map(fortworth_quant,'product_reference_code')
        apopka_map = df_to_qty_map(apopka_quant,'product_reference_code')
        winchester_map = df_to_qty_map(winchester_quant,'product_reference_code')
        miami_map = df_to_qty_map(miami_quant,'product_reference_code')
        bloomsburg_map = df_to_qty_map(bloomsburg_quant,'product_reference_code')
        tennesse_map = df_to_qty_map(tennesse_quant, 'part_number')

        product_map = {p['default_code']: p for p in products_info}
        product_map_pn = {p['part']: p for p in products_info}

        for row in records:
            product = product_map.get(row.get('PRODUCT'))
            product_pn = product_map_pn.get(row.get('PN'))

            full_size = product.get('full_size', '') if product else None
            tax_ids = product.get('taxes_id', []) if product else None
            
            row['BRAND'] = product.get('brand', '') if product else ''
            row['FULL_SIZE'] = product.get('full_size', '') if product else '0'
            row['TIRE_CLASS'] = product.get('tire_class', '') if product else ''
            row['PLY'] = product.get('ply', '') if product  else ''
            row['MODEL'] = product.get('model', '') if product else ''

            #For CAT_FILTER:
            cat_filer = product.get('lt_p', '') if product else ''
            try:
                
                patterns = ['LT', 'ST', 'X', 'C', r'\.5']

                match = next((p for p in patterns if re.search(p, full_size, re.I)), '')
                # Re-maps the CAT_FILTER for more accurate data
                mapping = {
                    'ST': 'ST',
                    'LT': 'LT',
                    'X': 'LT',
                    'C': 'VAN',
                    r'\.5': 'TBR',
                }

                # If CAT_Filter initially has a data, use it as is
                if cat_filer != '':
                    cat_filer = cat_filer
                    
                else:
                    cat_filer = mapping.get(match, 'P')
            except Exception as e:
                logger.warning(f"Error on CAT_FILTER for product {row.get('PRODUCT')}, Err msg: {e}")
                continue

            row['CAT_FILTER'] = cat_filer
            row['POSITION'] = product.get('tire_type', '') if product else ''
            row['FET'] = tax_map.get(tax_ids[0], 0) if tax_ids else 0
            row['PART'] = product.get('part', '') if product else ''

            #For strapping check
            weights = product.get('weight')
            if weights in ['0', None, '', 0]:
                weights = strap_map.get(full_size, '')
            
            row['WEIGHTS'] = weights
            row['SIZE'] = get_last_number(full_size)
            prod_code = row.get('PRODUCT')
            pn_code = row.get('PN')

            
            #Change this when gotten on hand
            row['POMONA'] = pm_map.get(prod_code, 0)
            row['CARROLTON'] = carr_map.get(prod_code, 0)
            row['LATROBE'] = latrobe_map.get(prod_code, 0) 
            row['FORT WORTH'] = fort_map.get(prod_code, 0)
            row['KENTUCKY'] = kent_map.get(prod_code, 0)
            row['APOPKA'] = apopka_map.get(prod_code, 0)
            row['WINCHESTER'] = winchester_map.get(prod_code, 0)
            row['MIAMI'] = miami_map.get(prod_code, 0)
            row['BLOOMSBURG'] = bloomsburg_map.get(prod_code, 0)
            row['TENNESSE'] = tennesse_map.get(pn_code, 0)

            
            row['LOAD_SPEED'] = (
                ''.join(filter(None, [
                    str(product.get('load_index')) if product and product.get('load_index') else '',
                    product.get('speed_rating') if product else ''
                ]))
                if product else ''
            )


            list_to_append_data.append(row)
        create_sheet(file, list_to_append_data, 'Sheet2')
        logger.info(f"Updated sheet with Odoo data, file: {os.path.basename(file)}")
    except Exception as e:
        logger.error(f"Error processing file '{file}': {e}")
        logger.exception(traceback.format_exc())
        return  


def populate_data(file, config, customer, capped):
    # context = ssl._create_unverified_context()
 
    # common = xmlrpc.client.ServerProxy(
    #     f"{ODOO_URL_B2B}/xmlrpc/2/common",
    #     context=context
    # )

    # uid = common.authenticate(ODOO_DB_B2B, ODOO_USERNAME_B2B, ODOO_PASSWORD_B2B, {})

    # if isinstance(uid, int) and uid > 0:
    #     logger.info(f"Odoo login as ID: {uid}")
    #     print('hello b2b')
    # else:
    #     logger.error("Failed")
    
    # models = xmlrpc.client.ServerProxy(f"{ODOO_URL_B2B}/xmlrpc/2/object", context=context)

    uid, models = get_models()

    #Added product identification step to filter products based on config before querying Odoo for pricelist items
    df = identify_prod_to_add(file, config, customer)
    #create_sheet(file, df.to_dict(orient= 'records'), customer)
    logger.info(f"Product identification completed, {type(df)} products to process for customer '{customer}'")
    codes = df['PRODUCT'].dropna().unique().tolist()

    # If customer is multiple price list 
    domain = [
        ('product_tmpl_id.default_code', 'in', codes),
        # ('fixed_price', '!=', 0),
        
    ] + (
        [('pricelist_id.name', 'ilike', customer)]
        if not isinstance(customer, list)
        else ['|'] * (len(customer) - 1) + [('pricelist_id.name', 'ilike', c) for c in customer]
    )



    products = models.execute_kw(
        ODOO_DB_B2B,
        uid,
        ODOO_PASSWORD_B2B,
        'product.pricelist.item',
        'search_read',
        [domain],
        {
            'fields': [
                'product_tmpl_id',
                'pricelist_id',
                'fixed_price',
                'write_date',
            ],
            'order': 'write_date desc',
        }
    )
    logger.info(f"Fetched:{products}")

    if not products:
        logger.warning("No pricelist items found for the given default_codes")
        return

    list_to_append_data = []
    
    price_index = defaultdict(list)
    
    
    # --------------------------------------------------
    # FILTER + INDEX (STRICT MATCH USING normalize)
    # --------------------------------------------------
    target = (
        [c.lower() for c in customer]
        if isinstance(customer, list)
        else customer.lower()
    )

    price_index = {}

    
    prod_df = pd.DataFrame([
        {
            "line_id": item["id"],
            "product_id": re.search(r'\[(.+)\]', item['product_tmpl_id'][1]).group(1) if re.search(r'\[(.+)\]', item['product_tmpl_id'][1]) else None,
            "pricelist_id": item["pricelist_id"][0],
            "pricelist_name": item["pricelist_id"][1],
            "fixed_price": item["fixed_price"],
            "write_date": pd.to_datetime(item["write_date"])
        }
        for item in products
    ]).to_dict(orient='records')

    for p in prod_df:
        pricelist = p.get('pricelist_name')
        product = p.get('product_id')

        # Safety checks
        if not pricelist or len(pricelist) < 2:
            continue
        if not product or len(product) < 2:
            continue

        pricelist_name = pricelist

        # Final strict check (cannot be done in Odoo domain)
        
        name_norm = normalize(pricelist_name)

        if isinstance(target, list):
            if not any(t in name_norm for t in target):
                continue
        else:
            if target not in name_norm:
                continue


        default_code = product

        # First record is latest (already ordered by write_date desc)
        if default_code not in price_index:
            price_index[default_code] = p

    logger.info(f"FINISHED INDEXING PRICES, total indexed products: {len(price_index)}")
    
    
    # --------------------------------------------------
    # BUILD FINAL OUTPUT
    # --------------------------------------------------
    list_to_append_data = []
    list_no_pricelist = []

    for _, row in df.iterrows():
        default_code = str(row['PRODUCT'])
        latest = price_index.get(default_code)
        brand = row['BRAND']
        type_ = row['TYPE']
        weight = row['WEIGHTS']
        ply = row['PLY']
        tire_class = row['TIRE_CLASS']
        full_size = row['FULL_SIZE']
        position = row['POSITION']
        model = row['MODEL']
        cat_filter = row['CAT_FILTER']
        load_speed = row['LOAD_SPEED'] 
        size = row['SIZE']
        if capped == 'Y':
            pomona = min(row['POMONA'], 200)
            carrolton = min(math.floor(row['CARROLTON'] * 0.75), 200)
            latrobe = min(row['LATROBE'], 200)
            fortw = min(row['FORT WORTH'], 200)
            kentucky = min(row['KENTUCKY'], 200)
            apopka = min(row['APOPKA'], 200)
            winchester = min(row['WINCHESTER'], 200)
            miami = min(row['MIAMI'], 200)
            bloomsburg = min(row['BLOOMSBURG'], 200)
            tennesse = min(row['TENNESSE'], 200)
        
        else:
            pomona = row['POMONA']
            carrolton = math.floor(row['CARROLTON'] * 0.75)
            latrobe = row['LATROBE']
            fortw = row['FORT WORTH']
            kentucky = row['KENTUCKY']
            apopka = row['APOPKA']
            winchester = row['WINCHESTER']
            miami = row['MIAMI']
            bloomsburg = row['BLOOMSBURG']
            tennesse = row['TENNESSE']
            
        fet = row['FET']
        part = row['PART']


    # Finding default codes without price rules
        if not latest:
            logger.warning(
                f"No {customer} price rules found for default_code={default_code} not adding it to the output"
            )
            # list_no_pricelist.append({customer: default_code})

            continue
    

        pricelist_name = latest['pricelist_name']
        price = latest['fixed_price']

        list_to_append_data.append({
            'brand': brand,
            'default_code': part,
            'weights': weight,
            'full_size': full_size,
            'type_': type_,
            'tire_class': tire_class,
            'ply': ply,

            #pattern is model
            'model': model,
            'position' : position,
            'size': size,
            'cat_filter': cat_filter,
            'load_speed': load_speed,
            'pricelist_name': pricelist_name,
            'California': pomona,
            'Carrolton': carrolton, 
            'Latrobe': latrobe,
            'FortWorth': fortw,
            'Kentucky': kentucky,
            'Apopka': apopka,
            'Winchester': winchester,
            'Miami': miami,
            'Bloomsburg': bloomsburg,
            'Tennesse':tennesse,
            'price': price,
            'fet' :fet,
        })

    # --------------------------------------------------
    # RESULT
    # --------------------------------------------------
    # list_to_append_data now contains final prices
    logger.info(f'Finished inputting {customer} pricelists')

    return_df = pd.DataFrame(list_to_append_data)
    if config['LOC_CODE'] in ['V']:
        warehouse_cols = ['California']

    elif config['LOC_CODE'] in ['PS']:
        warehouse_cols = ['California', 'Carrolton', 'Latrobe', 'FortWorth', 'Kentucky', 'Apopka']
    
    else:
        warehouse_cols = []

        list_loc_code = list(config['LOC_CODE'].upper())

        for loc in list_loc_code:
            if loc == 'V':
                warehouse_cols.append('California')
            elif loc == 'C':
                warehouse_cols.append('Carrolton')
            elif loc == 'L':
                warehouse_cols.append('Latrobe')
            elif loc == 'F':
                warehouse_cols.append('FortWorth')
            elif loc == 'K':
                warehouse_cols.append('Kentucky')
            elif loc == 'A':
                warehouse_cols.append('Apopka')
            elif loc == 'W':
                warehouse_cols.append('Winchester')
            elif loc == 'M':
                warehouse_cols.append('Miami')
            elif loc == 'B':
                warehouse_cols.append('Bloomsburg')
            elif loc == 'T':
                warehouse_cols.append('Tennesse')

    #Filtering data that are 8 and up inventory
    filter_df = return_df[(return_df[warehouse_cols] >= 8).any(axis=1)]

    # create_sheet(f'no_pricelist/data_no_pricelist_5.xlsx', list_no_pricelist, f'{customer}')


    return filter_df
# for data in config['Customer Name']:
#     populate_data('output/price_update.xlsx', data)
#populate_data(product_file, 'G_TOWNFAIR')


# def try_odoo():
#     warehouse = models.execute_kw(
#         ODOO_DB_B2B,
#         uid,
#         ODOO_PASSWORD_B2B,
#         'stock.quant',
#         'search_read',
#         [[('default_code', '=', 'AMERICANROADSTARARS4K02')]],
#         {
#             'fields': [
#                 'warehouse_id', 
#             ],
#         }
#     )
#     logger.info(f"Fetched:{warehouse}")



def copy_template(source, target, header_rows=5):
    # Column widths
    
    for col, dim in source.column_dimensions.items():
        target.column_dimensions[col].width = dim.width

    # Row heights
    for row, dim in source.row_dimensions.items():
        target.row_dimensions[row].height = dim.height

    # Merged cells
    for merged in source.merged_cells.ranges:
        target.merge_cells(str(merged))

    # Copy cells
    for row in source.iter_rows():
        for cell in row:
            tgt = target.cell(row=cell.row, column=cell.column)

            # ✅ Copy value ONLY if:
            # - Header row
            # - NOT a merged-cell placeholder
            if (
                cell.row <= header_rows
                and not isinstance(cell, MergedCell)
            ):
                tgt.value = cell.value

            # ✅ Styles are safe everywhere
            if cell.has_style:
                tgt.font = copy(cell.font)
                tgt.border = copy(cell.border)
                tgt.fill = copy(cell.fill)
                tgt.number_format = cell.number_format
                tgt.protection = copy(cell.protection)
                tgt.alignment = copy(cell.alignment)

#Hide columns
def hide_columns(file, column_names, ws_name:str = None):
    try:
        wb = load_workbook(file)

        column_to_hide = column_names

        if ws_name:
            if isinstance(ws_name, list):
                for name in ws_name:
                    ws = wb[name]

                    for col in column_to_hide:
                        ws.column_dimensions[col].hidden = True
            else:
                ws = wb[ws_name]
                
                for col in column_to_hide:
                    ws.column_dimensions[col].hidden = True

        else:
            for ws in wb.worksheets:
                for col in column_to_hide:
                    ws.column_dimensions[col].hidden = True
        
        wb.save(file)
        logger.info(f'Finished hiding columns')
    except Exception as e:
        logger.warning(f'Error on hiding columns, Err msg: {e}')

def unhide_columns(file, column_names, ws_name:str = None):
    try:
        wb = load_workbook(file)

        column_to_hide = column_names

        if ws_name:
            if isinstance(ws_name, list):
                for name in ws_name:
                    ws = wb[name]

                    for col in column_to_hide:
                        ws.column_dimensions[col].hidden = False
            else:
                ws = wb[ws_name]
                
                for col in column_to_hide:
                    ws.column_dimensions[col].hidden = False

        else:
            for ws in wb.worksheets:
                for col in column_to_hide:
                    ws.column_dimensions[col].hidden = False
        
        wb.save(file)
        logger.info(f'Finished hiding columns')
    except Exception as e:
        logger.warning(f'Error on hiding columns, Err msg: {e}')

def protect_sheet(file, cols_to_unlock):
    wb = load_workbook(file)

    for ws in wb.worksheets: 
    #Protect sheet (still blocks unhide)
        ws.protection.sheet = True
        ws.protection.formatColumns = True
    # ✅ Unlock ALL cells
        for row in ws.iter_rows(min_row=1, max_row=5):
            for cell in row:
                cell.protection = Protection(locked=False)
        for row in ws.iter_rows(min_row=6):
            for cell in row:
                cell.protection = Protection(locked=True)
        
        for cols in cols_to_unlock:
            for cell in ws[cols]:
                cell.protection = Protection(locked=False)    
        
    wb.save(file)

# Hide columns by range of columns
def hide_columns_range(file, start_col, end_col):
    wb = load_workbook(file)

    start = column_index_from_string(start_col)
    end = column_index_from_string(end_col)
    for ws in wb.worksheets:
        for i in range(start, end + 1):
            ws.column_dimensions[get_column_letter(i)].hidden = True
    
    wb.save(file)


def apply_currency_format(file, columns_names: list[str], ws_name=None):
    try:
        wb = load_workbook(file)
        columns_to_apply = columns_names
        currency_format = '"$"#,##0.00'

        if ws_name:
            if isinstance(ws_name, list):
                for name in ws_name:
                    ws = wb[name]
                    for col in columns_to_apply:
                        for cell in ws[col][5:]:
                                try:
                                    if cell.value is None:
                                        continue
                                    cell.value = float(cell.value)
                                    cell.number_format = currency_format
                                except (ValueError, TypeError):
                                    continue

            else:
                ws = wb[ws_name]
                for col in columns_to_apply:
                    for cell in ws[col][5:]:
                            try:
                                if cell.value is None:
                                    continue
                                cell.value = float(cell.value)
                                cell.number_format = currency_format
                            except (ValueError, TypeError):
                                continue
                
                
        else:
            for ws in wb.worksheets:
                for col in columns_to_apply:
                    for cell in ws[col][5:]:
                            try:
                                if cell.value is None:
                                    continue
                                cell.value = float(cell.value)
                                cell.number_format = currency_format
                            except (ValueError, TypeError):
                                continue
        logger.info(f'Applied currency on columns {columns_names}')
    
    
    
    except Exception as e:
        logger.warning(f'Error {file}, applying currency format: {e}')
    
    wb.save(file)


def delete_first_sheet(file, sheetname = 'Sheet1'):
    wb = load_workbook(file)
    ws= wb[sheetname]
    wb.remove(ws)
    wb.save(file)



def fill_rows(ws, row, fill, fill_type: int, location, ps_type = None):
    thin = Side(border_style='thin', color='000000')

    border = Border(left = thin, right = thin, top = thin, bottom = thin)

    types = ['norm','extra']
    conf_type = types[fill_type]
    order_fill = PatternFill(
        start_color= '86cb6b',
        end_color= '86cb6b',
        fill_type='solid'
    )

    if conf_type == 'extra':
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = fill
            ws.cell(row=row, column=col).border = border
    
    elif conf_type == 'norm':
        if location == 'V':
            for col in range(1, ws.max_column + 1):
                if col == column_index_from_string('K'):
                    ws.cell(row = row, column = col ).fill = order_fill
                else:
                    ws.cell(row=row, column=col).fill = fill

                ws.cell(row=row, column=col).border = border
        elif location == 'VLK':
            #Columns for Green Column
            column_names = ['K', 'S', 'AE', 'AN', 'AW', 'BB', 'BE', 'CP']
            column_indexes = [column_index_from_string(c) for c in column_names]
            for col in range(1, ws.max_column + 1):
                if col in column_indexes:
                    ws.cell(row = row, column = col ).fill = order_fill
                else:
                    ws.cell(row=row, column=col).fill = fill
                
                ws.cell(row=row, column=col).border = border

        elif location == 'PS':
            if ps_type == 'LTL':
                column_names = ['M', 'W', 'AH', 'AO']
            if ps_type == 'TL':
                column_names = ['J', 'T', 'AE' ,'AP']
            
            column_indexes = [column_index_from_string(c) for c in column_names]
            for col in range(1, ws.max_column + 1):
                if col in column_indexes:
                    ws.cell(row = row, column = col ).fill = order_fill
                else:
                    ws.cell(row=row, column=col).fill = fill
            
                ws.cell(row=row, column=col).border = border
        else:
            #Columns for Green Column
            column_names = ['K', 'S', 'AE', 'AN', 'AW', 'BF', 'BO', 'BX','CG']
            column_indexes = [column_index_from_string(c) for c in column_names]
            for col in range(1, ws.max_column + 1):
                if col in column_indexes:
                    ws.cell(row = row, column = col ).fill = order_fill
                else:
                    ws.cell(row=row, column=col).fill = fill
                
                ws.cell(row=row, column=col).border = border


def add_data(file, config, customer, hide_pricing = None, capped='Y'):
    loc_code = config['LOC_CODE']

    cust_df = populate_data(file, config, customer, capped)
    cust_df.columns = cust_df.columns.str.strip().str.upper()



    cust_df['SIZE_NUM'] = pd.to_numeric(cust_df['SIZE'], errors='coerce')
    logger.info(f'Len of Cust DF{len(cust_df)}')

    logger.info(f'Columns for customer: {cust_df.columns}')
    currency_format = '"$"#,##0.00'
    try:
        
        highlight_fill = PatternFill(
            start_color="d9d9d9",
            end_color="d9d9d9",
            fill_type="solid"
        )
        normal_fill = PatternFill(
            start_color="BBDDFC",
            end_color="BBDDFC",
            fill_type="solid"
        )
        yellow_fill = PatternFill(
            start_color="ffff99",
            end_color="ffff99",
            fill_type="solid"
        )
        white_fill= PatternFill(
            start_color="ffffff",
            end_color="ffffff",
            fill_type="solid"
        )

        if loc_code == 'V':
            clean_file = "format_excel/V_clean.xlsx"
            logger.info(f'Customer {customer}, using clean file: {clean_file}')
            
            wb = load_workbook(clean_file)
            template_ws = wb.active
            start_row = 6
            start_col =1

            #First Sheet
            ws_size = wb.create_sheet(title='By Size')
            copy_template(template_ws, ws_size)
            filtered_bysize = cust_df.sort_values(by='SIZE_NUM', ascending=True).reset_index(drop=True)
            logger.info(f'Len of filtered_bysize {len(filtered_bysize)}')
            
            prev_size = None
            current_row = start_row

            for r, row in filtered_bysize.iterrows():
                
                if prev_size is not None and row['SIZE_NUM'] != prev_size:
                    ws_size.insert_rows(current_row)
                    fill_rows(ws_size, current_row, highlight_fill, 1, loc_code)
                    current_row += 1  # move to the new row after insertion
                    logger.info(f'Appending row prev size: {prev_size}, {r}')

                for col_letter, col_config in mapping_excel.V_col_map.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    ws_size.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(ws_size, current_row, row_fill, 0, loc_code)


                prev_size = row['SIZE_NUM']
                current_row += 1

            logger.info(f'Added {customer} by size data to excel file')

            #----------------------------------------------------------------------------------------------------------------#

            #New Sheet
            ws_brand = wb.create_sheet(title='By Brand')
            #copying template
            copy_template(template_ws, ws_brand)
            cust_df['BRAND'] = cust_df['BRAND'].apply(normalize).str.upper()
            filtered_bybrand = cust_df.sort_values(by='BRAND', ascending=True).reset_index(drop=True)
            logger.info(f'Len of filtered_bybrand {len(filtered_bybrand)}')

            current_row = start_row
            prev_brand = None 

            for r, row in filtered_bybrand.iterrows():
                if prev_brand is not None and row['BRAND'] != prev_brand:
                    ws_brand.insert_rows(current_row)
                    fill_rows(ws_brand, current_row, highlight_fill, 1, loc_code)
                    
                    current_row += 1

                for col_letter, col_config in mapping_excel.V_col_map.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    ws_brand.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(ws_brand, current_row, row_fill, 0,loc_code)

                prev_brand = row['BRAND']
                current_row += 1

            logger.info(f'Added {customer} by brand data to excel file')

            result_file = f'result/{customer} California Specials {datetoday}.xlsx'
            
            wb.save(result_file)

            #---- Process extra formatting ----#

            hide_columns(result_file, ['P' , 'O'])
            apply_currency_format(result_file,['J','L', 'M'])

            if hide_pricing == 'Y':
                hide_columns(result_file, ['J', 'L', 'M'])
                logger.info(f'Hide pricing for customer: {customer}')
            delete_first_sheet(result_file)

        #----------------------------------------------------------------------------------------------------------------#
        #----------------------------------------------------------------------------------------------------------------#
        #----------------------------------------------------------------------------------------------------------------#
        
        #If all warehuose    
        elif loc_code == 'VLK':
            clean_file = 'format_excel/VLK_clean.xlsx'
            logger.info(f'Customer {customer}, using clean file: {clean_file}')
            
            wb = load_workbook(clean_file)
            template_ws = wb.active
            start_row = 6
            start_col =1

            #First Sheet
            ws_size = wb.create_sheet(title='By Size')
            copy_template(template_ws, ws_size)
            filtered_bysize = cust_df.sort_values(by='SIZE_NUM', ascending=True).reset_index(drop=True)
            logger.info(f'Len of filtered_bysize {len(filtered_bysize)}')

            prev_size = None
            current_row = start_row
           
            for r, row in filtered_bysize.iterrows():

                if prev_size is not None and row['SIZE_NUM'] != prev_size:
                    ws_size.insert_rows(current_row)
                    fill_rows(ws_size, current_row, highlight_fill, 1,loc_code)
                    current_row += 1  # move to the new row after insertion
                    logger.info(f'Appending row prev size: {prev_size}, {r}')

                
                for col_letter, col_config in mapping_excel.VLK_col_map.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    ws_size.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(ws_size, current_row, row_fill, 0,loc_code)

                prev_size = row['SIZE_NUM']
                current_row += 1


            logger.info(f'Added {customer} by size data to excel file')

            #----------------------------------------------------------------------------------------------------------------#
            
            #New Sheet
            ws_brand = wb.create_sheet(title='By Brand')
            #copying template
            copy_template(template_ws, ws_brand)
            cust_df['BRAND'] = cust_df['BRAND'].apply(normalize).str.upper()

            filtered_bybrand = cust_df.sort_values(by=['BRAND'],ascending=True).reset_index(drop=True)
            logger.info(f'Len of filtered_bybrand {len(filtered_bybrand)}')

            prev_brand = None

            current_row = start_row
            

            for r, row in filtered_bybrand.iterrows():

                if prev_brand is not None and row['BRAND'] != prev_brand:
                    ws_brand.insert_rows(current_row)
                    fill_rows(ws_brand, current_row, highlight_fill, 1, loc_code)
                    current_row += 1
                
                for col_letter, col_config in mapping_excel.VLK_col_map.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    ws_brand.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )
                
                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(ws_brand, current_row, row_fill, 0, loc_code)
                prev_brand = row['BRAND']
                current_row += 1
                
                

            result_file = f'result/{customer} California, Pennsylvania, Kentucky, Carrolton, Apopka and Fort Worth Specials {datetoday}.xlsx'
            logger.info(f'Added {customer} by brand data to excel file')
            
            wb.save(result_file)

            #---- Process extra formatting ----#
            hide_columns(result_file, ['BB', 'O', 'X', 'AI', 'AR', 'Y', 'AK' ,'AT', 'Z', 'AA','AB', 'BB','BJ','BK','BL','BM', 'BN', 'BO', 'BP', 'BQ','BR', 'BS', 'BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC'])
            apply_currency_format(result_file, ['J','L','M','R','T','U','AD','AF','AG','AM','AO','AP','AV','AX','AY'])
            if hide_pricing == 'Y':
                hide_columns(result_file, ['J','L','M','R','T','U','AD','AF','AG','AM','AO','AP','AV','AX','AY'])
                logger.info(f'Hide pricing for customer: {customer}')
            delete_first_sheet(result_file)
        
        #----------------------------------------------------------------------------------------------------------------#
        #----------------------------------------------------------------------------------------------------------------#
        #----------------------------------------------------------------------------------------------------------------#

        #For Point S
        elif loc_code == 'PS':
            clean_file = r'format_excel/PointS.xlsx'
            logger.info(f'Customer {customer}, using clean file: {clean_file}')

            wb = load_workbook(clean_file)
            LTL_template_ws = wb['LTL']
            LTL_ws_size = wb.create_sheet(title='LTL_By Size')
            copy_template(LTL_template_ws, LTL_ws_size)
            start_row = 6

            filtered_bysize = cust_df.sort_values(by='SIZE_NUM', ascending=True).reset_index(drop=True)
            
            prev_size = None

            for r, row in filtered_bysize.iterrows():
                
                current_row = start_row + r

                for col_letter, col_config in mapping_excel.ps_LTL.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    LTL_ws_size.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(LTL_ws_size, current_row, row_fill, 0, loc_code, 'LTL')

                                
                if prev_size is not None and row['SIZE_NUM'] != prev_size:
                    LTL_ws_size.insert_rows(current_row)
                    fill_rows(LTL_ws_size, current_row, highlight_fill, 1, loc_code)
                    current_row += 1  # move to the new row after insertion
                    logger.info(f'Appending row prev size: {prev_size}, {r}')
                        

                prev_size = row['SIZE_NUM']

            logger.info(f'Added {customer} by size data to excel file')

            #----------------------------------------------------------------------------------------------------------------#

            #New Sheet
            LTL_ws_brand = wb.create_sheet(title='LTL_By Brand')
            #copying template
            copy_template(LTL_template_ws, LTL_ws_brand)
            cust_df['BRAND'] = cust_df['BRAND'].apply(normalize).str.upper()
            filtered_bybrand = cust_df.sort_values(by='BRAND', ascending=True).reset_index(drop=True)

            current_row = start_row
            prev_brand = None 

            for r, row in filtered_bybrand.iterrows():
                if prev_brand is not None and row['BRAND'] != prev_brand:
                    LTL_ws_brand.insert_rows(current_row)
                    fill_rows(LTL_ws_brand, current_row, highlight_fill, 1, loc_code)
                    current_row += 1

                for col_letter, col_config in mapping_excel.ps_LTL.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    LTL_ws_brand.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(LTL_ws_brand, current_row, row_fill, 0,loc_code, 'LTL')

                prev_brand = row['BRAND']
                current_row += 1


            logger.info(f'Added {customer} by brand data to excel file')

            result_file = f'result/{customer} Specials {datetoday}.xlsx'
            wb.save(result_file)

            #---- Process extra formatting ----#
            
            

            #-----------------------------------------------------------#
            #-----------------------------------------------------------#

            TL_template_ws = wb['Trailer_Loads']
            TL_ws_size = wb.create_sheet(title='TL_By Size')
            copy_template(TL_template_ws, TL_ws_size)
            start_row = 6

            filtered_bysize = cust_df.sort_values(by='SIZE_NUM', ascending=True).reset_index(drop=True)
            
            prev_size = None

            for r, row in filtered_bysize.iterrows():
                
                current_row = start_row + r

                for col_letter, col_config in mapping_excel.ps_TL.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    TL_ws_size.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(TL_ws_size, current_row, row_fill, 0, loc_code, 'TL')

                                
                if prev_size is not None and row['SIZE_NUM'] != prev_size:
                    TL_ws_size.insert_rows(current_row)
                    fill_rows(TL_ws_size, current_row, highlight_fill, 1, loc_code)
                    current_row += 1  # move to the new row after insertion
                    logger.info(f'Appending row prev size: {prev_size}, {r}')
                        

                prev_size = row['SIZE_NUM']

            logger.info(f'Added {customer} by size data to excel file')

            #-----------------------------------------------------------------------------------------------#

            TL_ws_brand = wb.create_sheet(title='TL_By Brand')
            #copying template
            copy_template(TL_template_ws, TL_ws_brand)
            cust_df['BRAND'] = cust_df['BRAND'].apply(normalize).str.upper()
            filtered_bybrand = cust_df.sort_values(by='BRAND', ascending=True).reset_index(drop=True)

            current_row = start_row
            prev_brand = None 

            for r, row in filtered_bybrand.iterrows():
                if prev_brand is not None and row['BRAND'] != prev_brand:
                    TL_ws_brand.insert_rows(current_row)
                    fill_rows(TL_ws_brand, current_row, highlight_fill, 1, loc_code)
                    current_row += 1

                for col_letter, col_config in mapping_excel.ps_TL.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    TL_ws_brand.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(TL_ws_brand, current_row, row_fill, 0,loc_code, 'TL')

                prev_brand = row['BRAND']
                current_row += 1


            logger.info(f'Added {customer} by brand data to excel file')

            wb.save(result_file)

            #---- Process extra formatting ----#
            
            hide_columns(result_file, ['L','M','N','X','Y','Z','AA','AI','AJ','AK','AL','BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','AU','AV','AW','AX','AY','AZ','BL','BM','BN','BO','BP','BQ','BR', 'BS','BT','BU','BV','BW','BX','BY','BZ','CA','CB','CC'], ['TL_By Size', 'TL_By Brand'])
            apply_currency_format(result_file, ['I','K','O','S','U','V','AD','AF','AG','AO','AQ','AR'], ['TL_By Size', 'TL_By Brand'])

            hide_columns(result_file, ['I','J','K','S', 'T', 'U', 'AB', 'AD','AE','AF'], ['LTL_By Size', 'LTL_By Brand'])
            apply_currency_format(result_file, ['L','N','O','V','X','Y','AG','AI','AJ','AN','AP','AQ'], ['LTL_By Size', 'LTL_By Brand'])
            delete_first_sheet(result_file, 'LTL')
            delete_first_sheet(result_file, 'Trailer_Loads')
        
        else:
            logger.info(f'Customer using a dynamic Location Code:')
            clean_file = 'format_excel/Ware_Gen_clean.xlsx'
            logger.info(f'Customer {customer}, using clean file: {clean_file}')
            logger.info(f'Cust df Col: {cust_df.columns}')
            
            #These are the columns that will be unhidden in the final excel file
            col_to_unhide = []
            location_list_str = []

            #Location Code into a list to be consolidate columns to be unhidden
            list_loc_code = list(loc_code.upper())

            for loc in list_loc_code:
                if loc == 'V':
                    col_to_unhide.extend(['I','J','K','L','M'])
                    location_list_str.append("California")
                if loc == 'L':
                    col_to_unhide.extend(['P','Q','R','S','T','U','V'])
                    location_list_str.append("Pennsylvania")
                if loc == 'K':
                    col_to_unhide.extend(['Y','AC','AD','AE','AF','AG','AH'])
                    location_list_str.append("Kentucky")
                if loc == 'C':
                    col_to_unhide.extend(['AK','AL','AM','AN','AO','AP','AQ'])
                    location_list_str.append("Carrolton") 
                if loc == 'F':
                    col_to_unhide.extend(['AT','AU','AV','AW','AX','AY','AZ','BA'])
                    location_list_str.append("Fort Worth")
                if loc == 'A':
                    col_to_unhide.extend(['BC','BD','BE','BF','BG','BH','BI','BJ'])
                    location_list_str.append("Apopka")
                if loc == 'W':
                    col_to_unhide.extend(['BL','BM','BN','BO','BP','BQ', 'BR','BS'])
                    location_list_str.append("Winchester")
                if loc == 'M':
                    col_to_unhide.extend(['BU','BV','BW','BX','BY','BZ','CA','CB'])
                    location_list_str.append("Miami")
                if loc == 'B':
                    col_to_unhide.extend(['CD','CE','CF','CG','CH','CI','CJ','CK'])
                    location_list_str.append("Bloomsburg")
                if loc == 'T':
                    col_to_unhide.extend(['CN','CO','CP','CQ','CR','CS','CT'])
                    location_list_str.append("Tennesse")
                

            # This is in order to make location dynamic in the file name
            location_str = ', '.join(location_list_str)
            logger.info(f'Columns to unhide: {col_to_unhide}')
            


            wb = load_workbook(clean_file)
            template_ws = wb.active
            start_row = 6
            start_col =1

            #First Sheet
            ws_size = wb.create_sheet(title='By Size')
            copy_template(template_ws, ws_size)
            filtered_bysize = cust_df.sort_values(by='SIZE_NUM', ascending=True).reset_index(drop=True)
            logger.info(f'Len of filtered_bysize {len(filtered_bysize)}')

            prev_size = None
            current_row = start_row
           
            for r, row in filtered_bysize.iterrows():

                if prev_size is not None and row['SIZE_NUM'] != prev_size:
                    ws_size.insert_rows(current_row)
                    fill_rows(ws_size, current_row, highlight_fill, 1,loc_code)
                    current_row += 1  # move to the new row after insertion
                    logger.info(f'Appending row prev size: {prev_size}, {r}')

                
                for col_letter, col_config in mapping_excel.Ware_gen_col_map.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    ws_size.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )

                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(ws_size, current_row, row_fill, 0,loc_code)

                prev_size = row['SIZE_NUM']
                current_row += 1


            logger.info(f'Added {customer} by size data to excel file')

            #----------------------------------------------------------------------------------------------------------------#
            
            #New Sheet
            ws_brand = wb.create_sheet(title='By Brand')
            #copying template
            copy_template(template_ws, ws_brand)
            cust_df['BRAND'] = cust_df['BRAND'].apply(normalize).str.upper()

            filtered_bybrand = cust_df.sort_values(by=['BRAND'],ascending=True).reset_index(drop=True)
            logger.info(f'Len of filtered_bybrand {len(filtered_bybrand)}')

            prev_brand = None

            current_row = start_row
            

            for r, row in filtered_bybrand.iterrows():

                if prev_brand is not None and row['BRAND'] != prev_brand:
                    ws_brand.insert_rows(current_row)
                    fill_rows(ws_brand, current_row, highlight_fill, 1, loc_code)
                    current_row += 1
                
                for col_letter, col_config in mapping_excel.Ware_gen_col_map.items():
                    if col_config['type'] == 'value':
                        value = row[col_config['field']]
                    
                    elif col_config['type'] == 'formula':
                        value = col_config['field'].format(map_cell=current_row)
                    
                    ws_brand.cell(
                        row = current_row,
                        column = column_index_from_string(col_letter),
                        value = value
                    )
                
                if current_row % 2 == 0:
                    row_fill = yellow_fill
                else:
                    row_fill = white_fill

                fill_rows(ws_brand, current_row, row_fill, 0, loc_code)
                prev_brand = row['BRAND']
                current_row += 1
                
                

            result_file = f'result/{customer} {location_str} Specials {datetoday}.xlsx'
            logger.info(f'Added {customer} by brand data to excel file')
            
            wb.save(result_file)

            #---- Process extra formatting ----#
            hide_columns_range(result_file, 'I', 'DA')
            unhide_columns(result_file, col_to_unhide)
            apply_currency_format(result_file, ['J','L','M','R','T','U','AD','AF','AG','AM','AO','AP','AV','AX','AY'])
            if hide_pricing == 'Y':
                hide_columns(result_file, ['J','L','M','R','T','U','AD','AF','AG','AM','AO','AP','AV','AX','AY'])
                logger.info(f'Hide pricing for customer: {customer}')
            delete_first_sheet(result_file)
            #For protecting sheet for customers to only edit not unhide columns
            protect_sheet(result_file, ['K','S','AE','AN','AW','BF','BO', 'BX', 'CG'])

        #----- Return result file -----#
        return result_file

    
 
    except Exception as e:
        logger.error(f"Error determining clean file for customer '{customer}': {e}")
        logger.exception(traceback.format_exc())
        return 

    



